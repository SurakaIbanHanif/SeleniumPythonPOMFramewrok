import openpyxl


def get_data_from_excel(path, sheet_name):
    wb = openpyxl.load_workbook(path)

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' does not exist. Available: {wb.sheetnames}")

    sheet = wb[sheet_name]

    rows = sheet.max_row
    cols = sheet.max_column

    data = []
    for r in range(2, rows + 1):  #
        row_data = []
        for c in range(1, cols + 1):
            val = sheet.cell(row=r, column=c).value
            if val is None:
                val = ""
            if isinstance(val, str):
                val = val.strip()
            row_data.append(val)
        data.append(tuple(row_data))

    return data
